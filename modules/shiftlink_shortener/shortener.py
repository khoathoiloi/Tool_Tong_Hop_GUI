# -*- coding: utf-8 -*-
import os
import sys
import re
import time
import json
import random
import openpyxl
from playwright.sync_api import sync_playwright

TARGET_URL = "https://shorten-link-swart.vercel.app/"

def extract_links_from_sheet(sheet):
    best_col = None
    max_links = 0
    for col in range(1, sheet.max_column + 1):
        link_count = 0
        for row in range(2, min(50, sheet.max_row + 1)):
            val = str(sheet.cell(row=row, column=col).value or "")
            if re.search(r'https?://[^\s]+', val):
                link_count += 1
        if link_count > max_links:
            max_links = link_count
            best_col = col

    if not best_col:
        return None, []

    links = []
    for r in range(2, sheet.max_row + 1):
        cell_val = str(sheet.cell(row=r, column=best_col).value or "")
        match = re.search(r'https?://[^\s]+', cell_val)
        if match:
            links.append((r, match.group(0), cell_val))
            
    return best_col, links

def generate_short_slug(original_url, attempt=0):
    slug_part = re.sub(r'^https?://[^/]+(?:/[^/]+)*/', '', original_url.rstrip('/'))
    target_len = random.randint(15, 21)
    
    collected_chars = []
    char_count = 0
    for ch in reversed(slug_part):
        collected_chars.append(ch)
        if ch != '-':
            char_count += 1
        if char_count >= target_len:
            break
            
    result = ''.join(reversed(collected_chars))
    result = re.sub(r'^[^a-zA-Z0-9]+', '', result)
    result = re.sub(r'[^a-zA-Z0-9]+$', '', result)
    
    if attempt > 1:
        suffix = ''.join(random.choices('abcdefghijklmnopqrstuvwxyz0123456789', k=2))
        result = f"{result[:18]}-{suffix}"
        
    return result

def wait_for_user_login(page, log_cb, stop_check_cb):
    time.sleep(1.5)
    user_profile = page.locator("#user-profile")
    user_display_name = page.locator("#user-display-name")
    profile_classes = user_profile.get_attribute("class") or ""
    
    if user_profile.is_visible() and "hidden" not in profile_classes:
        username = user_display_name.inner_text().strip() or "User"
        log_cb(f"👤 Tài khoản đã đăng nhập: {username}", "SUCCESS")
        return True

    log_cb("⚠️ CHƯA ĐĂNG NHẬP! Xin mời bạn đăng ký / đăng nhập trực tiếp trên Chrome...", "WARNING")
    log_cb("⏳ Đang đợi bạn đăng nhập...", "INFO")

    while not stop_check_cb():
        time.sleep(1)
        try:
            if page.is_closed():
                log_cb("❌ Trình duyệt Chrome đã bị đóng.", "ERROR")
                return False
            p_classes = user_profile.get_attribute("class") or ""
            if user_profile.is_visible() and "hidden" not in p_classes:
                username = user_display_name.inner_text().strip() or "User"
                log_cb(f"🎉 ĐĂNG NHẬP THÀNH CÔNG! Chào mừng {username}.", "SUCCESS")
                return True
        except Exception:
            pass
    return False

def fetch_domains_from_web(user_data_dir, log_cb, stop_check_cb):
    """Mở trình duyệt, chờ đăng nhập và lấy toàn bộ danh sách tên miền từ website"""
    os.makedirs(user_data_dir, exist_ok=True)
    domains = []
    
    with sync_playwright() as playwright:
        context = playwright.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            channel="chrome",
            headless=False,
            args=["--start-maximized"],
            no_viewport=True
        )
        page = context.pages[0] if context.pages else context.new_page()
        page.goto(TARGET_URL, wait_until="domcontentloaded")

        if not wait_for_user_login(page, log_cb, stop_check_cb):
            context.close()
            return []

        time.sleep(1.5)
        try:
            page.wait_for_selector("#domain-select option", timeout=5000)
        except Exception:
            pass

        raw_opts = page.eval_on_selector_all("#domain-select option", "elems => elems.map(e => e.innerText.trim()).filter(t => t.length > 0)")
        if raw_opts:
            domains = list(dict.fromkeys(raw_opts))
            log_cb(f"✅ Đã nạp thành công {len(domains)} tên miền từ website: {', '.join(domains)}", "SUCCESS")
        else:
            log_cb("⚠️ Chưa tìm thấy tên miền nào trong dropdown web.", "WARNING")

        context.close()
        
    return domains

def shorten_multiple_urls(raw_urls, selected_domain, show_browser, user_data_dir, log_cb, progress_cb, stop_check_cb):
    """
    Rút gọn một danh sách các URL gốc trực tiếp (dùng cho quy trình Combo Tạo Excel Fanpage)
    Trả về dict: {raw_url: shortened_url}
    """
    unique_urls = [u for u in dict.fromkeys(raw_urls) if u and u.startswith("http")]
    if not unique_urls:
        return {}

    log_cb(f"🚀 Bắt đầu rút gọn {len(unique_urls)} đường link qua ShiftLink (Tên miền: {selected_domain})...", "HIGHLIGHT")
    os.makedirs(user_data_dir, exist_ok=True)
    url_map = {}

    with sync_playwright() as playwright:
        context = playwright.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            channel="chrome",
            headless=not show_browser,
            args=["--start-maximized"] if show_browser else [],
            no_viewport=True if show_browser else False
        )
        page = context.pages[0] if context.pages else context.new_page()
        page.goto(TARGET_URL, wait_until="domcontentloaded")

        if not wait_for_user_login(page, log_cb, stop_check_cb):
            context.close()
            return {}

        time.sleep(1.0)
        total = len(unique_urls)

        for idx, original_url in enumerate(unique_urls, 1):
            if stop_check_cb():
                log_cb("⚠️ Đã dừng tiến trình rút gọn.", "WARNING")
                break

            if progress_cb:
                progress_cb(idx, total)

            success = False
            for attempt in range(5):
                if stop_check_cb(): break
                slug = generate_short_slug(original_url, attempt=attempt)

                page.locator("#original-url").fill(original_url)
                try:
                    page.locator("#domain-select").select_option(label=selected_domain)
                except Exception:
                    try:
                        page.locator("#domain-select").select_option(value=selected_domain)
                    except Exception:
                        pass

                page.locator("#short-path").fill(slug)
                time.sleep(0.3)

                try:
                    submit_btn = page.locator("#shorten-form button[type='submit']")
                    with page.expect_response(lambda res: "/api/links" in res.url and res.request.method == "POST", timeout=12000) as resp_info:
                        submit_btn.click()

                    response = resp_info.value
                    toast_msg = "Thành công"
                    try:
                        toast_elem = page.locator("#toast.show")
                        toast_elem.wait_for(state="visible", timeout=2500)
                        toast_msg = toast_elem.inner_text().strip()
                    except Exception:
                        pass

                    if response.status in [200, 201]:
                        res_data = response.json()
                        created_path = res_data.get("shortPath", slug)
                        shortened_url = f"https://{selected_domain}/{created_path}"
                        url_map[original_url] = shortened_url
                        log_cb(f"[{idx:02d}/{total:02d}] ✅ Đã rút gọn: {original_url} -> {shortened_url}", "SUCCESS")
                        success = True
                        time.sleep(1.2)
                        break
                    else:
                        log_cb(f"[{idx:02d}/{total:02d}] ⚠️ Thử lại slug khác ({toast_msg})...", "WARNING")
                        time.sleep(0.8)
                except Exception as e:
                    time.sleep(0.8)

            if not success and not stop_check_cb():
                log_cb(f"[{idx:02d}/{total:02d}] ❌ Không thể rút gọn link: {original_url}", "ERROR")

        context.close()

    return url_map

def run_shorten_automation(excel_path, selected_domain, show_browser, target_sheet_names, user_data_dir, log_cb, progress_cb, stop_check_cb):
    excel_path = excel_path.strip('\'"')
    if not os.path.exists(excel_path):
        log_cb(f"❌ Lỗi: Không tìm thấy file Excel tại '{excel_path}'", "ERROR")
        return False, None

    wb = openpyxl.load_workbook(excel_path)
    sheets_to_process = []
    if target_sheet_names:
        sheets_to_process = [s for s in target_sheet_names if s in wb.sheetnames]
    else:
        sheets_to_process = wb.sheetnames

    tasks = []
    total_links_all = 0

    for sname in sheets_to_process:
        sheet = wb[sname]
        col_idx, links = extract_links_from_sheet(sheet)
        if links:
            out_col = None
            for c in range(1, sheet.max_column + 1):
                if str(sheet.cell(1, c).value or '').strip().lower() == "link da rut gon":
                    out_col = c
                    break
            if not out_col:
                out_col = sheet.max_column + 1
                sheet.cell(row=1, column=out_col, value="Link da rut gon")
                
            tasks.append((sname, sheet, col_idx, out_col, links))
            total_links_all += len(links)

    if not tasks:
        log_cb("❌ Không tìm thấy đường link nào trong các sheet đã chọn!", "ERROR")
        return False, None

    log_cb(f"📊 Bắt đầu xử lý: {len(tasks)} Sheet với tổng cộng {total_links_all} dòng link.", "INFO")

    url_cache = {}
    processed_count = 0
    os.makedirs(user_data_dir, exist_ok=True)

    with sync_playwright() as playwright:
        context = playwright.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            channel="chrome",
            headless=not show_browser,
            args=["--start-maximized"] if show_browser else [],
            no_viewport=True if show_browser else False
        )
        page = context.pages[0] if context.pages else context.new_page()
        page.goto(TARGET_URL, wait_until="domcontentloaded")

        if not wait_for_user_login(page, log_cb, stop_check_cb):
            context.close()
            return False, None

        time.sleep(1.0)

        for sname, sheet, col_idx, out_col, items in tasks:
            log_cb(f"📂 Đang xử lý Sheet: '{sname}' ({len(items)} links)...", "HIGHLIGHT")
            for row_idx, original_url, full_cell_text in items:
                if stop_check_cb():
                    log_cb("⚠️ Đã dừng tiến trình rút gọn.", "WARNING")
                    break

                processed_count += 1
                progress_cb(processed_count, total_links_all)
                
                if original_url in url_cache:
                    cached_val = url_cache[original_url]
                    sheet.cell(row=row_idx, column=out_col, value=cached_val)
                    log_cb(f"[{processed_count:03d}/{total_links_all:03d}] [{sname}] Dòng #{row_idx:<3d} -> ⚡ [DÙNG LẠI] {cached_val}", "INFO")
                    continue

                success = False
                for attempt in range(5):
                    if stop_check_cb():
                        break
                    slug = generate_short_slug(original_url, attempt=attempt)
                    
                    page.locator("#original-url").fill(original_url)
                    try:
                        page.locator("#domain-select").select_option(label=selected_domain)
                    except Exception:
                        try:
                            page.locator("#domain-select").select_option(value=selected_domain)
                        except Exception:
                            pass

                    page.locator("#short-path").fill(slug)
                    time.sleep(0.3)

                    try:
                        submit_btn = page.locator("#shorten-form button[type='submit']")
                        with page.expect_response(lambda res: "/api/links" in res.url and res.request.method == "POST", timeout=12000) as resp_info:
                            submit_btn.click()

                        response = resp_info.value
                        toast_msg = "Thành công"
                        try:
                            toast_elem = page.locator("#toast.show")
                            toast_elem.wait_for(state="visible", timeout=2500)
                            toast_msg = toast_elem.inner_text().strip()
                        except Exception:
                            pass

                        if response.status in [200, 201]:
                            res_data = response.json()
                            created_path = res_data.get("shortPath", slug)
                            shortened_url = f"https://{selected_domain}/{created_path}"
                            final_text = f"watch full here 👉: {shortened_url}"
                            
                            url_cache[original_url] = final_text
                            sheet.cell(row=row_idx, column=out_col, value=final_text)
                            
                            log_cb(f"[{processed_count:03d}/{total_links_all:03d}] [{sname}] Dòng #{row_idx:<3d} -> ✅ {final_text}", "SUCCESS")
                            success = True
                            time.sleep(1.2)
                            break
                        else:
                            log_cb(f"[{processed_count:03d}/{total_links_all:03d}] [{sname}] Dòng #{row_idx:<3d} -> ⚠️ Thử lại slug khác ({toast_msg})...", "WARNING")
                            time.sleep(0.8)
                    except Exception as e:
                        time.sleep(0.8)

                if not success and not stop_check_cb():
                    log_cb(f"[{processed_count:03d}/{total_links_all:03d}] [{sname}] Dòng #{row_idx:<3d} -> ❌ Thất bại sau 5 lần thử.", "ERROR")

            if stop_check_cb():
                break

        context.close()

    dir_name = os.path.dirname(excel_path)
    base_name = os.path.splitext(os.path.basename(excel_path))[0]
    out_file = os.path.join(dir_name, f"{base_name}_shortened.xlsx")
    wb.save(out_file)
    return True, out_file